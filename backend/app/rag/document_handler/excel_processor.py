from typing import List, Dict, Tuple
from langchain_core.documents import Document


class ExcelProcessor:
    """Excel 处理器：行→自然语言，支持多级表头、合并单元格和多 sheet"""

    async def process(self, file_path: str) -> List[Document]:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        documents = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 构建合并单元格映射
            merged_map = self._build_merged_map(ws)

            # 检测多级表头
            header_rows, data_start_row = self._detect_headers(rows)
            headers = self._build_headers(rows, header_rows, merged_map)

            sheet_context = f"[Sheet: {sheet_name}]"

            for row_idx in range(data_start_row, len(rows) + 1):  # +1 因为行号从 1 开始
                row = rows[row_idx - 1]
                values = []
                for col_idx in range(len(row)):
                    val = self._get_cell_value(row, col_idx, row_idx, merged_map)
                    values.append(str(val) if val is not None else "")

                if not any(values):
                    continue

                parts = []
                for col_idx, v in enumerate(values):
                    if v and col_idx < len(headers):
                        parts.append(f"{headers[col_idx]}{v}")

                if parts:
                    nl_text = f"{sheet_context} " + "，".join(parts)
                    documents.append(Document(
                        page_content=nl_text,
                        metadata={
                            "source": file_path,
                            "sheet": sheet_name,
                            "row": row_idx,
                        }
                    ))

        return documents

    def _build_merged_map(self, ws) -> Dict[Tuple[int, int], object]:
        """构建合并单元格映射：(row, col) → 左上角单元格值"""
        merged_map = {}
        for merged_range in ws.merged_cells.ranges:
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            # 获取左上角单元格的值
            top_left = ws.cell(row=min_row, column=min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = top_left
        return merged_map

    def _get_cell_value(self, row: tuple, col_idx: int, row_num: int,
                        merged_map: Dict[Tuple[int, int], object]) -> object:
        """获取单元格值，合并单元格返回左上角值"""
        key = (row_num, col_idx + 1)  # openpyxl 列从 1 开始
        if key in merged_map:
            return merged_map[key]
        if col_idx < len(row):
            return row[col_idx]
        return None

    def _detect_headers(self, rows: list) -> Tuple[List[int], int]:
        """
        检测多级表头行，返回 (表头行号列表, 数据起始行号)
        启发式：
        - 第一行始终视为表头
        - 后续行如果全部单元格都是纯文本标签（不含数字、长度<30）
          且与上一行列数一致 → 识别为表头
        - 最多识别 3 级表头
        """
        header_rows = [1]
        max_scan = min(5, len(rows))

        for i in range(1, max_scan):
            row = rows[i]
            if not row:
                break
            # 过滤有效值
            str_vals = []
            for v in row:
                if v is None:
                    str_vals.append("")
                else:
                    s = str(v).strip()
                    str_vals.append(s)

            non_empty = [s for s in str_vals if s]
            if not non_empty:
                break

            # 数据行特征：包含数字、或文本较长（>50字）
            has_number = any(
                s.replace(".", "").replace("-", "").isdigit()
                for s in non_empty
            )
            has_long = any(len(s) > 50 for s in non_empty)

            # 与前一行列数一致（排除空列尾部）
            prev_cols = len([c for c in rows[i - 1] if c is not None and str(c).strip()]) if i > 0 else len(non_empty)

            is_label_row = (
                not has_number
                and not has_long
                and len(header_rows) < 3
                and len(non_empty) <= prev_cols + 2  # 允许少量列扩展
            )

            if is_label_row:
                header_rows.append(i + 1)
            else:
                break

        data_start = max(header_rows) + 1
        return header_rows, data_start

    def _build_headers(self, rows: list, header_rows: List[int],
                       merged_map: Dict[Tuple[int, int], object]) -> List[str]:
        """构建最终表头列表，多级用 '>' 连接"""
        max_cols = max(len(rows[r - 1]) if r <= len(rows) else 0 for r in header_rows)
        headers = []

        for col_idx in range(max_cols):
            parts = []
            for r in header_rows:
                if r <= len(rows):
                    row = rows[r - 1]
                    val = self._get_cell_value(row, col_idx, r, merged_map)
                    if val is not None and str(val).strip():
                        parts.append(str(val).strip())
            if parts:
                headers.append(' > '.join(parts))
            else:
                headers.append(f'列{col_idx + 1}')

        return headers
