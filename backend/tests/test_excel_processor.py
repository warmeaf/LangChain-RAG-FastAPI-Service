"""测试 1.2: Excel 合并单元格 + 多级表头"""

import pytest
import openpyxl
import tempfile
import os
from app.rag.document_handler.excel_processor import ExcelProcessor


class TestExcelProcessor:
    """企业级验收标准：合并单元格值填充 + 多级表头 '>' 连接"""

    def _create_test_xlsx(self, rows_data):
        """创建测试用 xlsx 并返回路径"""
        wb = openpyxl.Workbook()
        ws = wb.active
        for row_idx, row in enumerate(rows_data, start=1):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        return tmp.name

    def _create_merged_xlsx(self):
        """创建含合并单元格的测试 xlsx"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工信息"
        # 合并 A1:B1
        ws.merge_cells("A1:B1")
        ws["A1"] = "基本信息"
        ws["C1"] = "薪资"
        # 二级表头
        ws["A2"] = "姓名"
        ws["B2"] = "职位"
        ws["C2"] = "月薪"
        # 数据
        ws["A3"] = "张三"
        ws["B3"] = "工程师"
        ws["C3"] = 15000
        ws["A4"] = "李四"
        ws["B4"] = "经理"
        ws["C4"] = 25000
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        return tmp.name

    @pytest.mark.asyncio
    async def test_document_count(self):
        """文档数 = 数据行数"""
        path = self._create_merged_xlsx()
        try:
            docs = await ExcelProcessor().process(path)
            assert len(docs) == 2, f"期望 2 行数据，实际 {len(docs)}"
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_multi_level_headers(self):
        """多级表头：'基本信息 > 姓名张三'"""
        path = self._create_merged_xlsx()
        try:
            docs = await ExcelProcessor().process(path)
            flat = " ".join(d.page_content for d in docs)
            assert "基本信息 > 姓名张三" in flat, f"未找到多级表头格式: {flat[:200]}"
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_merged_cells_share_parent(self):
        """合并单元格：'姓名' 和 '职位' 共享父级 '基本信息'"""
        path = self._create_merged_xlsx()
        try:
            docs = await ExcelProcessor().process(path)
            flat = " ".join(d.page_content for d in docs)
            assert "基本信息 > 职位" in flat, f"职位应共享基本信息父级: {flat[:200]}"
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_sheet_context_prefix(self):
        """Sheet 上下文：以 '[Sheet: 员工信息]' 开头"""
        path = self._create_merged_xlsx()
        try:
            docs = await ExcelProcessor().process(path)
            for d in docs:
                assert d.page_content.startswith("[Sheet: 员工信息]"), \
                    f"缺少 Sheet 前缀: {d.page_content[:80]}"
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_handles_empty_rows(self):
        """空行被跳过"""
        path = self._create_test_xlsx([
            ["姓名", "年龄"],
            ["张三", 30],
            [None, None],  # 空行
            ["李四", 25],
        ])
        try:
            docs = await ExcelProcessor().process(path)
            assert len(docs) == 2, f"空行应跳过，实际 {len(docs)}"
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_simple_table_no_merge(self):
        """简单表格（无合并单元格）正常处理"""
        path = self._create_test_xlsx([
            ["产品", "价格"],
            ["苹果", 5],
            ["香蕉", 3],
        ])
        try:
            docs = await ExcelProcessor().process(path)
            assert len(docs) == 2
            flat = " ".join(d.page_content for d in docs)
            assert "产品苹果" in flat
        finally:
            os.unlink(path)
